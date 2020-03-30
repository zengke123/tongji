#!/usr/bin/env python
# 日报绘图
import os
import logging
import datetime
from util import db
from util.file_exists import get_date_range
from collections import namedtuple
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis


def logger():
    logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s [%(levelname)s] %(message)s',
                    datefmt='%Y-%m-%d %H:%M:%S')


class DrawChart():

    '''
    绘制曲线图
    title: 图表标题
    style: 颜色类型
    data: 数据
    dates: 日期
    '''

    def __init__(self, worksheet, title):
        self.chart = LineChart()
        self.ws = worksheet
        self.chart.title = title
        self.chart.y_axis.crossAx = 500
        self.chart.x_axis = DateAxis(crossAx=100)
        self.chart.x_axis.number_format = 'd-mmm'
        self.chart.x_axis.majorTimeUnit = "days"

    def set_x(self, x_title):
        # 设置横坐标标题
        self.chart.x_axis.title = x_title

    def set_y(self, y_titel):
        # 设置纵坐标标题
        self.chart.y_axis.title = y_titel

    def set_style(self, style):
        # 设置图标样式
        self.chart.style = style

    def add_data(self, datas, row, col):
        # 将数据写入sheet指定区域
        for i, data in enumerate(datas):
            for j, value in enumerate(data):
                self.ws.cell(row=row + i, column=col + j, value=value)
        # 全部数据范围
        self.data = Reference(self.ws, min_col=col+1, min_row=row, max_col=col+len(datas[1])-1, max_row=32)
        # 日期数据范围
        self.dates = Reference(self.ws, min_col=1, min_row=row+1, max_row=32)

    def draw(self):
        self.chart.add_data(self.data, titles_from_data=True)
        self.chart.set_categories(self.dates)
        return self.chart


def get_data(dates, clustes, data_type):
    row_title = clustes[:]
    row_title.insert(0,'时间')
    rows = []
    rows.append(row_title)
    for date in dates:
        row = []
        row.append(date)
        for cluste in clustes:
            sql = "select {} from as_pfmc where date={} and cluste='{}'".format(data_type, date, cluste)
            result = db.select(sql)
            try:
                row.append(result[0][0])
            except IndexError:
                row.append(None)
        rows.append(row)
    return rows


def get_chart(worksheet, title, style, data, row, col):
    c = DrawChart(worksheet=worksheet, title=title)
    c.set_style(style)
    c.add_data(datas=data, row=row, col=col)
    return c.draw()


def main():
    # 网元名称需要与网管上保持一致
    catas = ['CLAS04', 'CLAS05', 'CLAS06', 'sccl20', 'sccl22','CLAS03','CLAS07','CLAS07']
    scpas = ['scpas03', 'scpas04', 'scpas05', 'scpas06', 'scpas07', 'scpas08', 'scpas35', 'scpas38']
    sicp = ['SICP01', 'SICP02', 'SICP03', 'SICP04']
    smp = ['SMP2']
    vc = ['VC3', 'VC4']
    ctx = ['CTX02', 'CTX03', 'CTXAS04', 'CTX05']
    today = datetime.datetime.now().strftime("%Y%m%d")
    # 日志记录
    # logger()
    # 数据库连接
    db.create_engine('root', 'zengke', 'tongji', '127.0.0.1')
    # 取最近7天的数据
    date_range = get_date_range(today, delta_day=30)
    if os.path.exists("line.xlsx"):
        wb = load_workbook("line.xlsx", data_only=True)
        isRemove = True
    else:
        wb = load_workbook("templates/d_report.xlsx", data_only=True)
        isRemove = False
    ws = wb.get_sheet_by_name("原始数据")
    ws2 = wb.get_sheet_by_name("主机性能数据")
    ws3 = wb.get_sheet_by_name("主机性能数据2")
    # 集群类型，数据类型，标题，图表样式，excel原始数据行，数据列，图表放置位置
    AsChart = namedtuple("AsChart", "cluste_type data_type title style row col location")
    as_list = [AsChart(catas, "max_cpu", "CATAS CPU占用率(%)", 10, 2, 1, "A1"),
               AsChart(catas, "max_mem", "CATAS MEM占用率(%)", 10, 2, 11, "A18"),
               AsChart(catas, "max_io", "CATAS IO占用率(%)", 10, 2, 21, "A32"),
               AsChart(scpas, "max_cpu", "SCPAS CPU占用率(%)", 10, 2, 31, "K1"),
               AsChart(scpas, "max_mem", "SCPAS MEM占用率(%)", 10, 2, 41, "K18"),
               AsChart(scpas, "max_io", "SCPAS IO占用率(%)", 10, 2, 51, "K32"),
               AsChart(ctx, "max_cpu", "CTX CPU占用率(%)", 10, 2, 97, "U1"),
               AsChart(ctx, "max_mem", "CTX MEM占用率(%)", 10, 2, 102, "U18"),
               AsChart(ctx, "max_io", "CTX IO占用率(%)", 10, 2, 107, "U32")]
    as_list2 = [AsChart(sicp, "max_cpu", "SICP CPU占用率(%)", 10, 2, 61, "A1"),
               AsChart(sicp, "max_mem", "SICP MEM占用率(%)", 10, 2, 67, "A18"),
               AsChart(sicp, "max_io", "SICP IO占用率(%)", 10, 2, 73, "A32"),
               AsChart(smp, "max_cpu", "SMP CPU占用率(%)", 10, 2, 79, "K1"),
               AsChart(smp, "max_mem", "SMP MEM占用率(%)", 10, 2, 82, "K18"),
               AsChart(smp, "max_io", "SMP IO占用率(%)", 10, 2, 85, "K32"),
               AsChart(vc, "max_cpu", "VC CPU占用率(%)", 10, 2, 88, "U1"),
               AsChart(vc, "max_mem", "VC MEM占用率(%)", 10, 2, 91, "U18"),
               AsChart(vc, "max_io", "VC IO占用率(%)", 10, 2, 94, "U32")]
    # 生成CATAS图表
    for chart in as_list:
        as_data_rows = get_data(date_range, chart.cluste_type, chart.data_type)
        as_chart = get_chart(ws, chart.title, chart.style, data=as_data_rows, row=chart.row, col=chart.col)
        ws2.add_chart(as_chart, chart.location)
    for chart in as_list2:
        as_data_rows = get_data(date_range, chart.cluste_type, chart.data_type)
        as_chart = get_chart(ws, chart.title, chart.style, data=as_data_rows, row=chart.row, col=chart.col)
        ws3.add_chart(as_chart, chart.location)
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    filename = "d_report_" + str(today) + ".xlsx"
    wb.save(filename)
    if isRemove:
        os.remove("line.xlsx")

if __name__ == "__main__":
    db.create_engine('root', 'zengke', 'tongji', '127.0.0.1')
    main()
