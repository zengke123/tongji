import config
from datetime import datetime
from util.file_exists import file_exists, get_data
from openpyxl import load_workbook


def get_dict(filename):
    caps = get_data(filename)
    caps_dict = {}
    for x in caps:
        try:
            x.remove('')
        except ValueError:
            pass
        if len(x) == 2:
            caps_dict[x[0]] = float(x[1])
        else:
            caps_dict[x[0]] = x[1:]
    return caps_dict


@file_exists(config.src_files.get('ypfmc'))
def get_pfmc(filename):
    return get_dict(filename)


@file_exists(config.src_files.get('tsucc'))
def get_tsucc(filename):
    return get_dict(filename)


@file_exists(config.src_files.get('tsucc_y'))
def get_tsucc_y(filename):
    return get_dict(filename)


@file_exists(config.src_files.get('tcaps'))
def get_tcaps(filename):
    return get_dict(filename)


@file_exists(config.src_files.get('tcaps_y'))
def get_tcaps_y(filename):
    return get_dict(filename)


def write_to_excel(ws, row, col, datas):
    for i, values in enumerate(datas):
        for j, value in enumerate(values):
            ws.cell(row=row + i, column=col + j, value=value)


def main():
    wb = load_workbook("templates/load.xlsx")
    ws = wb.get_sheet_by_name("Sheet1")
    # 从文件中提取数据
    pfmc = get_pfmc()
    tsucc = get_tsucc()
    tsucc_y = get_tsucc_y()
    tcaps = get_tcaps()
    tcaps_y = get_tcaps_y()

    # 按指定格式获取数据，cluste为网元名称
    def parse_data(cluste, pre_name):
        cluste_t = pre_name + cluste
        _t_caps = tcaps.get(cluste_t, 0)
        _y_caps = tcaps_y.get(cluste_t, 0)
        _c_caps = config.caps_capacity.get(cluste_t, 0)
        _row = [cluste, float(pfmc.get(cluste)[0]), 100 - float(pfmc.get(cluste)[1]), _t_caps, _y_caps, _c_caps,
                format(_t_caps / _c_caps, '.2%')]
        if _y_caps and _y_caps != 0:
            _row.append(format((_t_caps - _y_caps) / _y_caps, '.2%'))
        else:
            _row.append(0)
        return _row

    # 获取SCPAS数据
    scpas_data = []
    for item in config.cluters_dict.get('SCPAS'):
        row = parse_data(item, pre_name='SCP-')
        scpas_data.append(row)
    # 获取音频彩铃数据
    crbt_data = []
    for item in ["CLAS03", "CLAS04", "CLAS05", "CLAS06", "sccl20", "sccl22"]:
        row = parse_data(item, pre_name='CRBT-')
        crbt_data.append(row)
    # 获取视频彩铃数据
    vrbt_data = []
    for item in ["CLAS03", "CLAS07"]:
        if item == "CLAS03":
            t_caps = tcaps.get("CAVTAS03")
            y_caps = tcaps_y.get("CAVTAS03")
            c_caps = config.caps_capacity.get('CRBT-CAVTAS03')
            row = [item, float(pfmc.get(item)[0]), 100 - float(pfmc.get(item)[1]), t_caps, y_caps,
                   c_caps, format(t_caps / c_caps, '.2%'), format((t_caps - y_caps) / y_caps, '.2%')]
            vrbt_data.append(row)
        if item == "CLAS07":
            t_caps = tcaps.get("CAVTAS07")
            y_caps = tcaps_y.get("CAVTAS07")
            c_caps = config.caps_capacity.get('CRBT-CLAS07')
            row = [item, float(pfmc.get(item)[0]), 100 - float(pfmc.get(item)[1]), t_caps, y_caps,
                   c_caps, format(t_caps / c_caps, '.2%'), format((t_caps - y_caps) / y_caps, '.2%')]
            vrbt_data.append(row)
    # 获取centrex数据
    centrex_data = []
    for item in config.cluters_dict.get('CENTREX'):
        row = parse_data(item, pre_name='Centrex-')
        centrex_data.append(row)
    # 获取SCP数据
    scp_data = []
    for item in config.cluters_dict.get('SCP'):
        row = parse_data(item, pre_name='SCP-')
        scp_data.append(row)
    # 获取SCOM数据
    scom_data = []
    for item in config.cluters_dict.get('SICP'):
        row = [item, float(pfmc.get(item)[0]), 100 - float(pfmc.get(item)[1])]
        scom_data.append(row)
    # 获取2g 彩铃数据
    cl_data = []
    for item in config.cluters_dict.get('CL'):
        item_t = 'CRBT-' + item
        row = [item, float(pfmc.get(item)[0]), 100 - float(pfmc.get(item)[1]), tsucc.get(item_t), tsucc_y.get(item_t)]
        cl_data.append(row)
    # 填写excel
    write_to_excel(ws, row=4, col=1, datas=scpas_data)
    write_to_excel(ws, row=14, col=1, datas=crbt_data)
    write_to_excel(ws, row=23, col=1, datas=vrbt_data)
    write_to_excel(ws, row=28, col=1, datas=centrex_data)
    write_to_excel(ws, row=32, col=1, datas=scp_data)
    write_to_excel(ws, row=57, col=1, datas=scom_data)
    write_to_excel(ws, row=63, col=1, datas=cl_data)
    date = datetime.now().strftime("%Y%m%d")
    outfile = "load" + date + ".xlsx"
    wb.save(outfile)


if __name__ == "__main__":
    main()
