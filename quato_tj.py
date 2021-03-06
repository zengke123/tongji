#!/usr/bin/env python
import os, shutil
import datetime
import logging
from openpyxl import load_workbook
from util.TxtParse import parseHtml
from modules.get_caps import get_caps
from modules.get_ywzb import get_ywzb
from modules.get_streamnumber import get_streamnumber, get_scpas_streamnumber
from modules.get_users import get_v_users
from modules.get_cpu_html import cpu_analyse
from modules.get_week_caps import get_week_caps
from modules.get_dailycaps import wirte_caps_to_excle
from modules.get_vsstj import get_vsstj
from modules.get_cpu_data import wirte_cpu_to_excle
# 2018-07-26 :增加ctx_caps指标
# 2018-08-03 :增加视频彩铃指标
# 2018-10-10 :增加SCPAS话单流水号
# 2018-10-16 :调整业务指标
# 2018-10-23 :业务指标数据入库ywzb
# 2018-11-01 :日报增加caps统计
# 2018-11-20 :增加短号短信数据入库
# 2020-03-30 :调整模板，CPU数据字获取

# 填写业务指标到日报
def write_to_report(ywzb_dict, max_cluster, max_streamnumber):
    # 将指标填入日报
    wb = load_workbook("templates/d_report.xlsx", data_only=True)
    ws = wb.get_sheet_by_name("用户数&关键指标")
    # 2/3G V网呼叫成功率 2gscp_minsucc
    scp_minsucc = ywzb_dict.get('2gscp_minsucc', '')
    if isinstance(scp_minsucc, list):
        ws.cell(row=45, column=2, value=str(scp_minsucc[1])+'%')
        ws.cell(row=45, column=3, value=scp_minsucc[0])
    # 2/3G 彩铃播放成功率 2gcl_minsucc
    cl_minsucc = ywzb_dict.get('2gcl_minsucc', '')
    if isinstance(cl_minsucc, list):
        ws.cell(row=46, column=2, value=str(cl_minsucc[1])+'%')
        ws.cell(row=46, column=3, value=cl_minsucc[0])
    # SCP忙时CAPS数 2gscp_maxcaps
    scp_maxcaps = ywzb_dict.get('2gscp_maxcaps', '')
    if isinstance(scp_maxcaps, list):
        ws.cell(row=47, column=2, value=scp_maxcaps[1])
        ws.cell(row=47, column=3, value=scp_maxcaps[0])
    # SCP最大话单流水号
    ws.cell(row=48, column=2, value=max_streamnumber)
    ws.cell(row=48, column=3, value=max_cluster)
    # 二卡充值成功率 UCCminsucc
    UCCminsucc = ywzb_dict.get('UCCminsucc', '')
    if isinstance(UCCminsucc, list):
        ws.cell(row=49, column=2, value=str(UCCminsucc[1])+'%')
        ws.cell(row=49, column=3, value=UCCminsucc[0])
    # SCPAS网络接通率 SCPAS_minnetsucc
    SCPAS_minnetsucc = ywzb_dict.get('SCPAS_minnetsucc', '')
    if isinstance(SCPAS_minnetsucc, list):
        ws.cell(row=39, column=2, value=str(SCPAS_minnetsucc[1])+'%')
        ws.cell(row=39, column=3, value=SCPAS_minnetsucc[0])
    # 彩铃AS网络接通率 CLAS_minnetsucc
    CLAS_minnetsucc = ywzb_dict.get('CLAS_minnetsucc', '')
    if isinstance(CLAS_minnetsucc, list):
        ws.cell(row=41, column=2, value=str(CLAS_minnetsucc[1])+'%')
        ws.cell(row=41, column=3, value=CLAS_minnetsucc[0])
    # 彩铃AS invite响应率 CLAS_mininvite
    CLAS_mininvite = ywzb_dict.get('CLAS_mininvite', '')
    if isinstance(CLAS_mininvite, list):
        ws.cell(row=42, column=2, value=str(CLAS_mininvite[1])+'%')
        ws.cell(row=42, column=3, value=CLAS_mininvite[0])
    # SCPAS invite响应率 CLAS_mininvite
    SCPAS_mininvite = ywzb_dict.get('SCPAS_mininvite', '')
    if isinstance(SCPAS_mininvite, list):
        ws.cell(row=40, column=2, value=str(SCPAS_mininvite[1])+'%')
        ws.cell(row=40, column=3, value=SCPAS_mininvite[0])
    # 彩铃AS 放音成功率 CLAS_minplaysucc
    CLAS_minplaysucc = ywzb_dict.get('CLAS_minplaysucc', '')
    if isinstance(CLAS_minplaysucc, list):
        ws.cell(row=43, column=2, value=str(CLAS_minplaysucc[1])+'%')
        ws.cell(row=43, column=3, value=CLAS_minplaysucc[0])
    # 视频彩铃AS 放音成功率 CLAS_minplaysucc
    VRBT_minplaysucc = ywzb_dict.get('VRBT_minplaysucc', '')
    if isinstance(VRBT_minplaysucc, list):
        ws.cell(row=44, column=2, value=str(VRBT_minplaysucc[1])+'%')
        ws.cell(row=44, column=3, value=VRBT_minplaysucc[0])
    # 提取cluters中对应的网元的caps数据填入日报(33,6)位置，新增网元在cluters增加即可
    cluters = ['SCP-scpas03', 'SCP-scpas04', 'SCP-scpas05', 'SCP-scpas06', 'SCP-scpas07', 'SCP-scpas08',
               'SCP-scpas35', 'SCP-scpas38', 'SCP-scp27', 'CRBT-CLAS03', 'CRBT-CLAS07', 'CRBT-CLAS08']
    wirte_caps_to_excle(cluters=cluters, ws=ws, row=36, col=6)
    wirte_cpu_to_excle(cluters=cluters, ws=ws, row=36, col=10)
    wb.save('line.xlsx')


# 业务指标汇总
def quato_analyse():
    quato_data = [["指标项", "集群", "指标"]]
    streamnumber_result = get_streamnumber()
    if streamnumber_result:
        max_cluster, max_streamnumber = streamnumber_result
    else:
        max_cluster, max_streamnumber = "", ""
    scpas_streamnumber_result = get_scpas_streamnumber()
    if scpas_streamnumber_result:
        scpas_max_cluster, scpas_max_streamnumber = scpas_streamnumber_result
    else:
        scpas_max_cluster, scpas_max_streamnumber = "", ""
    ywzb_result = get_ywzb()
    if ywzb_result:
        ywzb_data, ywzb_dict = ywzb_result
        quato_data.extend(ywzb_data)
        # 将指标填入日报
        write_to_report(ywzb_dict, max_cluster, max_streamnumber)

    # 汇总最大话单流水号
    quato_data.extend([["SCP最大话单流水号", max_cluster, max_streamnumber],
                      ["SCPAS最大话单流水号", scpas_max_cluster, scpas_max_streamnumber]])
    # 汇总用户数统计
    users = get_v_users()
    quato_data.extend([["V网Volte用户数","SCPAS", users.get('vpmn_volte')],
                       ["音频彩铃用户数", "CATAS", users.get('crbt_volte')],
                       ["视频彩铃用户数", "CATAS", users.get('vrbt')]])
    # 汇总caps统计
    caps_data = get_caps()
    if caps_data:
        quato_data.extend(caps_data)
    # 生成邮件正文html
    quato_html = parseHtml(quato_data, title="关键业务指标", return_all=False)
    return quato_html


def main():
    today = datetime.date.today()
    quato_report = "report_maxcpu_" + str(today) + ".html"
    if os.path.exists(quato_report):
        os.remove(quato_report)
    shutil.copy("templates/alarm.html", quato_report)
    t_start = '<table class="tableizer-table" cellspacing=0 width="60%";>\n'
    t_end = '</table>\n'
    zyjh_html, cpu_html = "", ""
    try:
        cpu_analyse_result = cpu_analyse()
        if cpu_analyse_result:
            zyjh_html, cpu_html = cpu_analyse_result
    except Exception as e:
        logging.error(e)
        logging.error("作业计划指标统计生成失败")
    quato_html = quato_analyse()
    # 短号短信数据入库
    get_vsstj()
    # 周四生成一周caps 数据
    if today.weekday() == 3:
        caps = get_week_caps()
        caps_html = t_start + caps + t_end
    else:
        caps_html = ""
    html = t_start + quato_html + t_end + "<br></br>" + caps_html + "<br></br>" +\
           t_start + cpu_html + t_end + "<br></br>" + t_start + zyjh_html + t_end
    with open(quato_report, "a") as f:
        f.write(html)
    logging.info("业务指标统计生成成功")


if __name__ == '__main__':
    main()



